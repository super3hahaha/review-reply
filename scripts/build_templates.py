"""
build_templates.py — 把 data/source/*.xlsx 编译成 data/templates.json。

xlsx 结构（每个产品一个 sheet）：
  A 列 = 类别中文名（可能为空 → 继承上一行）
  B 列 = 英文模板（主语种，总是有）
  C/D, E/F, ... = 其他语言译文（**忽略不读**，运行时由 skill 翻译）

输出（data/templates.json）：
  {
    "version": "<source mtime>",
    "source_file": "...",
    "products": {
      "XFolder": {
        "templates": [
          {"id": "xfolder-001", "category": "要五星", "text": "Hi there, thanks..."},
          ...
        ]
      },
      ...
    }
  }

调用：python scripts/build_templates.py
约定：空 B 列（无英文）→ 跳过；空 A 列继承上一行类别。
"""
from __future__ import annotations
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "data" / "source"
OUT_FILE = ROOT / "data" / "templates.json"
# 紧凑「匹配索引」：每条模板只留 id + category（不含全文）。批量回复（路径 A）匹配阶段
# 只读这个 → 上下文极小、思考量小。命中后再按 id 从 templates.json 取全文翻译。
OUT_INDEX = ROOT / "data" / "index.json"

# sheet 名 → 产品规范名（package_map.json 用的也是这套）
SHEET_TO_PRODUCT = {
    "XFolder": "XFolder",
    "MP3 cutter": "MP3 Cutter",
    "video to MP3": "Video to MP3",
}

PRODUCT_SLUGS = {
    "XFolder": "xfolder",
    "MP3 Cutter": "mp3cutter",
    "Video to MP3": "video2mp3",
}


def clean(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def slugify_category(s: str) -> str:
    """类别名 → 安全 id slug 片段（保留中文，去标点空格）"""
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[\\/\:\*\?\"\<\>\|\.,;]", "", s)
    return s or "uncategorized"


def parse_sheet(ws, product: str) -> list[dict]:
    """逐行扫 sheet，只读 A（类别）和 B（英文模板）。空 B → 跳过；空 A → 继承上一类别。"""
    slug = PRODUCT_SLUGS[product]
    templates: list[dict] = []
    current_category = ""
    counter = 0

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        category = clean(row[0]) if len(row) > 0 else ""
        english = clean(row[1]) if len(row) > 1 else ""

        if category:
            current_category = category

        if not english:
            continue

        counter += 1
        templates.append(
            {
                "id": f"{slug}-{counter:03d}",
                "category": current_category or "未分类",
                "text": english,
            }
        )

    return templates


def main():
    sources = sorted(SOURCE_DIR.glob("*.xlsx"))
    sources = [p for p in sources if not p.name.startswith("~$")]
    if not sources:
        raise SystemExit(f"在 {SOURCE_DIR} 下找不到 xlsx 源文件")
    if len(sources) > 1:
        print(f"警告：找到多个 xlsx，使用第一个: {sources[0].name}", file=sys.stderr)
    src = sources[0]
    print(f"读取: {src}")

    wb = load_workbook(src, data_only=True, read_only=True)

    warnings: list[str] = []
    products: dict[str, dict] = {}
    for sheet_name, product in SHEET_TO_PRODUCT.items():
        if sheet_name not in wb.sheetnames:
            warnings.append(f"找不到 sheet '{sheet_name}'，跳过产品 '{product}'")
            continue
        tpls = parse_sheet(wb[sheet_name], product)
        products[product] = {"templates": tpls}
        print(f"  {product}: {len(tpls)} 条")

    out = {
        "version": datetime.fromtimestamp(src.stat().st_mtime, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_file": src.name,
        "products": products,
    }

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n写入: {OUT_FILE}  ({OUT_FILE.stat().st_size} bytes)")

    # ——— 额外产出紧凑「匹配索引」 index.json（全部模板，仅 id + category）———
    idx_products: dict[str, dict] = {}
    for product, pdata in products.items():
        idx_products[product] = {
            "templates": [{"id": t["id"], "category": t["category"]} for t in pdata["templates"]]
        }
    idx_out = {
        "version": out["version"],
        "source_file": src.name,
        "note": "匹配索引：全部模板的 id+category（无全文）。路径 A 匹配阶段只读这个；命中后按 id 从 templates.json 取全文翻译。",
        "products": idx_products,
    }
    OUT_INDEX.write_text(json.dumps(idx_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n写入: {OUT_INDEX}  ({OUT_INDEX.stat().st_size} bytes)")

    if warnings:
        print(f"\n{len(warnings)} 条警告：", file=sys.stderr)
        for w in warnings:
            print(f"  - {w}", file=sys.stderr)


if __name__ == "__main__":
    main()

"""
copy_shared_templates.py — 把指定类别的模板从一个 sheet **物理复制**到另一个 sheet。

背景：某些 app（如 Video to MP3）希望拥有 MP3 Cutter 的部分类别模板的**独立副本**
（不是共用引用 —— 复制后两边可各自独立编辑）。本脚本只复制 A 列（类别）+ B 列（英文），
不碰 C 列起的翻译公式列（build 本就忽略它们）。

幂等：目标 sheet 已存在同名类别则跳过，可安全重跑。

改完后必须重跑 build_templates.py 重新生成 templates.json。

用法：python scripts/copy_shared_templates.py
"""
from __future__ import annotations
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "data" / "source"

# (源 sheet, 目标 sheet, 要复制的类别名列表)
COPY_JOBS = [
    (
        "MP3 cutter",
        "video to MP3",
        [
            "说好但没有五星",
            "无意义评论求五星",
            "无意义差评",
            "缺点+建议没给五星",
            "询问详细问题",
            "询问详细建议",
            "广告太多",
            "免费使用",
            "广告加载失败",
            "有害广告",
            "截图或录屏",
            "打不开文件",
            "裁剪后音质降低",
            "裁剪不准确1",
            "裁剪不准确2",
            "播放器精度不显示0.1s的问题",
        ],
    ),
]


def clean(v) -> str:
    return "" if v is None else str(v).strip()


def parse_blocks(ws) -> "dict[str, list[str]]":
    """sheet → {类别: [英文模板, ...]}，遵守"空 A 继承上一类别"。"""
    out: dict[str, list[str]] = {}
    cur = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a = clean(row[0]) if len(row) > 0 else ""
        b = clean(row[1]) if len(row) > 1 else ""
        if a:
            cur = a
            out.setdefault(cur, [])
        if b and cur is not None:
            out[cur].append(b)
    return out


def existing_categories(ws) -> set:
    cats = set()
    for row in ws.iter_rows(values_only=True):
        if row and len(row) > 0 and row[0] is not None:
            c = str(row[0]).strip()
            if c:
                cats.add(c)
    return cats


def last_data_row(ws) -> int:
    """最后一个 A 或 B 列非空的行号（避免 max_row 被空白格式行抬高）。"""
    last = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row:
            continue
        a = clean(row[0]) if len(row) > 0 else ""
        b = clean(row[1]) if len(row) > 1 else ""
        if a or b:
            last = i
    return last


def main():
    sources = sorted(p for p in SOURCE_DIR.glob("*.xlsx") if not p.name.startswith("~$"))
    if not sources:
        raise SystemExit(f"在 {SOURCE_DIR} 下找不到 xlsx 源文件")
    src = sources[0]
    print(f"打开: {src}")
    wb = load_workbook(src)  # data_only=False：保留现有公式（缓存值会在保存时丢失，build 不读译文列，无影响）

    total_added = 0
    for src_sheet, dst_sheet, categories in COPY_JOBS:
        if src_sheet not in wb.sheetnames:
            print(f"  跳过：源 sheet '{src_sheet}' 不存在", file=sys.stderr)
            continue
        if dst_sheet not in wb.sheetnames:
            print(f"  跳过：目标 sheet '{dst_sheet}' 不存在", file=sys.stderr)
            continue
        src_blocks = parse_blocks(wb[src_sheet])
        dst_ws = wb[dst_sheet]
        dst_cats = existing_categories(dst_ws)
        row_cursor = last_data_row(dst_ws)
        print(f"\n{src_sheet} → {dst_sheet}（目标当前末行 {row_cursor}）")
        for cat in categories:
            if cat not in src_blocks or not src_blocks[cat]:
                print(f"  ⚠ 源中无类别或无英文模板，跳过：{cat}", file=sys.stderr)
                continue
            if cat in dst_cats:
                print(f"  · 目标已存在，跳过：{cat}")
                continue
            texts = src_blocks[cat]
            for j, text in enumerate(texts):
                row_cursor += 1
                # 第一行写类别名，后续变体行 A 留空（沿用源约定）
                dst_ws.cell(row=row_cursor, column=1, value=cat if j == 0 else None)
                dst_ws.cell(row=row_cursor, column=2, value=text)
                total_added += 1
            dst_cats.add(cat)
            print(f"  ✓ 复制 {len(texts)} 条：{cat}")

    if total_added == 0:
        print("\n没有新增任何模板（可能都已存在）。未改动文件。")
        return

    wb.save(src)
    print(f"\n已保存，新增 {total_added} 条模板到 {src.name}")
    print("→ 现在请重跑：python scripts/build_templates.py")


if __name__ == "__main__":
    main()
